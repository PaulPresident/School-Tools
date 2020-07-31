# from openpyxl import Workbook
from openpyxl import load_workbook


wb = load_workbook(filename='Retport Card Template - Copy.xlsx')
ws = wb.active

ws['C1'] = 'Paul Sedra TEST TEST TEST TEST TEST TEST TEST nfoishgoishfoiahifjaiof'


wb.save('Retport Card Template - Copy.xlsx')
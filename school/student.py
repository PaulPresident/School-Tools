from openpyxl import load_workbook

from school.year import Year
from school.report_card import ReportCard
from school.transcript import Transcript

class Student():
    def __init__(self, name:str):
        self.name = name
        self._year = 0      # Grade
        self._mp = 0        # current
        self.high_school = {
            9: Year(9),         10: Year(10),
            11: Year(11),       12: Year(12),
        }

    @property
    def mp(self):
        return self._mp

    @mp.setter
    def mp(self, mp:int):
        if not isinstance(mp, int): raise TypeError('Please Provide an Integer!')
        if mp > 4: raise ValueError('Invalid Marking Period: Marking Period cannot exceed 4!')
        self._mp = mp

    @property
    def year(self):
        return self._year

    @year.setter
    def year(self, year:int):
        if not isinstance(year, int): raise TypeError('Please Provide an Integer!')
        if not 9 <= year <= 12: raise ValueError('Invlid Year: Only Grades 9-12 are supported!')
        self._year = year

    def report_card(self, year:int, mp:int):
        if (year ** (year%8) * mp) > (self._year ** (self._year%8) * self._mp):
            raise ValueError('Cannot request a Report Card for a Year or Marking Period higher than your last given!')
        return ReportCard(mp=mp, year=self.high_school.get(year))

    def write_report_card(self, year:int, mp:int):
        rc = self.report_card(year=year, mp=mp)
        wb = rc.write()
        ws = wb.active
        ws['C1'] = self.name
        ws['J1'] = year

        wb.save(f'{self.name} Report Card {year}.{mp}.xlsx')

    def transcript(self, year:int):
        if year > self._year:
            raise ValueError('Cannot request a Transcript of a year higher than your last given!')
        return Transcript(years=[year_obj for grade, year_obj in self.high_school.items() if grade <= year])

    def write_transcript(self, year:int):
        ts = self.transcript(year=year)
        wb = ts.write()
        ws = wb.active
        ws['A2'] = self.name
        ws['E2'] = year

        wb.save(f'{self.name} Transcript {year}.xlsx')
from openpyxl import load_workbook
from school.year import Year

class ReportCard():
    ROW = 6
    def __init__(self, mp:int, year:Year):
        self._mp = mp
        self._year = year
        self._subjects = [subject for subject in self._year.subjects]
        self._subjects_in_mp = [
            subject
            for subject in self._subjects
            if subject.mp[self._mp].grade._grade
        ]
        self._grades = [
            subject.mp[self._mp].grade
            for subject in self._subjects_in_mp
        ]
        self._credits = sum(subject.credit for subject in self._subjects_in_mp)

    @property
    def unweighted_nga(self):
        grades = sum(grade.unweighted for grade in self._grades)
        return round(grades / len(self._subjects_in_mp), 3)

    @property
    def weighted_nga(self):
        grades = sum(grade.weighted for grade in self._grades)
        return round(grades / self._credits, 3)

    @property
    def unweighted_gpa(self):
        gpas = sum(grade.unweighted_gpa for grade in self._grades)
        return round(gpas / len(self._subjects_in_mp), 3)

    @property
    def weighted_gpa(self):
        gpas = sum(grade.weighted_gpa for grade in self._grades)
        return round(gpas / len(self._subjects_in_mp), 3)

# TODO no subject lower than 80 or 70 in marking period and test
    @property
    def honor_roll(self):
        if self.unweighted_nga >= 93:
            return 'First Honor Roll'
        if self.unweighted_nga >= 83:
            return 'Second Honor Roll'



    def _write_name_and_credits(self, ws, subject):
        self.ROW += 1
        ws.cell(column=1, row=self.ROW, value=subject.name)
        ws.cell(column=12, row=self.ROW, value=f'{subject.credit:.3f}')

    def _write_grades(self, ws, subject):
        for mp, col in {1:3, 2:4, 3:7, 4:8}.items():
            try:
                ws.cell(column=col, row=self.ROW, value=subject.mp[mp].grade.unweighted)
            except TypeError:
                continue

# TODO refine
# I don't like these if's and try's research?
    def _write_exams_final_grades(self, ws, subject):
        # columns = [5, 9, 6, 10, 11]
        # values = [subject.sem1.exam, subject.sem2.exam, subject.sem1.final, subject.sem2.final, subject.final.unweighted]
        # generator = map(lambda col, value: ws.cell(column=col, row=self.ROW, value=value), columns, values)

        # for iteration in generator:
        #     try:
        #         next(iteration)
        #     except TypeError:
        #         continue


        if subject.sem1.exam:
            ws.cell(column=5, row=self.ROW, value=subject.sem1.exam)
        if subject.sem2.exam:
            ws.cell(column=9, row=self.ROW, value=subject.sem2.exam)
        try:
            ws.cell(column=6, row=self.ROW, value=subject.sem1.final)
        except TypeError: pass
        try:
            ws.cell(column=10, row=self.ROW, value=subject.sem2.final)
        except TypeError: pass
        try:
            ws.cell(column=11, row=self.ROW, value=subject.final.unweighted)
        except TypeError: pass

    def write(self):
        wb = load_workbook(filename='Retport Card Template.xlsx')
        ws = wb.active

        for subject in self._subjects:
            self._write_name_and_credits(ws=ws, subject=subject)
            self._write_grades(ws=ws, subject=subject)
            self._write_exams_final_grades(ws=ws, subject=subject)

        ws['A23'] = f'Congratulations on achieving {self.honor_roll}.'
        ws['A24'] = f'MP{self._mp} NGA: {self.weighted_nga}'

        return wb

    def _test_write(self):
        wb = self.write()
        wb.save('Test Report Card.xlsx')
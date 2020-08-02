from openpyxl import load_workbook
import itertools
from typing import List
from school.year import Year

class Transcript():
    def __init__(self, years:List[Year]):
        self._years = years
        self._subjects = [
            subject
            for year in self._years
            for subject in year.subjects
        ]

    @property
    def credits(self):
        return sum(
            year.credits
            for year in self._years
        )

    @property
    def credits_earned(self):
        return sum(
            year.credits_earned
            for year in self._years
        )

    @property
    def unweighted_nga(self):
        ngas = [subject.final.unweighted for subject in self._subjects]
        return round(sum(ngas) / len(self._subjects), 3)

    @property
    def weighted_nga(self):
        ngas = [subject.final.weighted for subject in self._subjects]
        return round(sum(ngas) / self.credits, 3)

# TODO fix GPA calculation
# NOTE Calculation of GPA is incorrect might go back to Grade Object
    @property
    def unweighted_gpa(self):
        gpas = [subject.final.unweighted_gpa for subject in self._subjects]
        return round(sum(gpas) / len(self._subjects), 3)

    @property
    def weighted_gpa(self):
        gpas = [subject.final.weighted_gpa for subject in self._subjects]
        return round(sum(gpas) / len(self._subjects), 3)



    def _write_year(self, ws, year, row_1):
        row = row_1
        if not year.completed: return
        for subject in year.subjects:
            ws.cell(column=1, row=row, value=subject.name)
            ws.cell(column=4, row=row, value=f'{subject.final.unweighted}')
            ws.cell(column=5, row=row, value=f'{subject.credit:.2f}')
            row += 1

        ws[f'A{row_1+10}'] = f'Crd Att: {year.credits:.3f}'
        ws[f'B{row_1+10}'] = f'CMP: {year.credits_earned:.3f}'
        ws[f'D{row_1+10}'] = f'NGA: {year.weighted_nga:.3f}'

    def _write_nga_summary(self, ws):
        rows = [5, 6, 7, 8, 10, 11]
        values = [self.weighted_nga, self.unweighted_nga, self.weighted_gpa, self.unweighted_gpa, self.credits, self.credits_earned]
        list(map(lambda row, value: ws.cell(column=11, row=row, value=f'{value:.3f}'), rows, values))

    def write(self):
        wb = load_workbook(filename='Transcript Template.xlsx')
        ws = wb.active

        list(map(self._write_year, itertools.repeat(ws), self._years, [6, 19, 32, 45]))
        self._write_nga_summary(ws=ws)

        return wb

    def _test_write(self):
        wb = self.write()
        wb.save('Test Transcript.xlsx')